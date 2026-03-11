"""
Broad Reach Logistics – Analysis Excel Generator
Produces a branded, multi-tab workbook for client download.
"""

import io, json, math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XlImage

# ── Brand palette ──────────────────────────────────────────────────────────────
NAVY       = "1B2A4A"
TEAL       = "0D9488"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F8FAFC"
MID_GRAY   = "E2E8F0"
DARK_TEXT   = "1E293B"
MUTED_TEXT  = "64748B"
GREEN_BG   = "ECFDF5"
GREEN_TEXT  = "047857"
RED_BG     = "FEF2F2"
RED_TEXT    = "DC2626"
GOLD_BG    = "FFFBEB"
GOLD_TEXT   = "D97706"
TEAL_LIGHT = "F0FDFA"

# Style presets
FONT_TITLE   = Font(name="Aptos", size=18, bold=True, color=WHITE)
FONT_H2      = Font(name="Aptos", size=13, bold=True, color=NAVY)
FONT_H3      = Font(name="Aptos", size=11, bold=True, color=NAVY)
FONT_HEADER  = Font(name="Aptos", size=10, bold=True, color=WHITE)
FONT_BODY    = Font(name="Aptos", size=10, color=DARK_TEXT)
FONT_BODY_SM = Font(name="Aptos", size=9, color=MUTED_TEXT)
FONT_KPI_VAL = Font(name="Aptos", size=22, bold=True, color=NAVY)
FONT_KPI_LBL = Font(name="Aptos", size=9, color=MUTED_TEXT)
FONT_GREEN   = Font(name="Aptos", size=10, bold=True, color=GREEN_TEXT)
FONT_RED     = Font(name="Aptos", size=10, bold=True, color=RED_TEXT)
FONT_LINK    = Font(name="Aptos", size=10, color="0066CC", underline="single")

FILL_NAVY    = PatternFill("solid", fgColor=NAVY)
FILL_TEAL    = PatternFill("solid", fgColor=TEAL)
FILL_WHITE   = PatternFill("solid", fgColor=WHITE)
FILL_ALT     = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_HEADER  = PatternFill("solid", fgColor=NAVY)
FILL_GREEN   = PatternFill("solid", fgColor=GREEN_BG)
FILL_RED     = PatternFill("solid", fgColor=RED_BG)
FILL_GOLD    = PatternFill("solid", fgColor=GOLD_BG)
FILL_TEAL_LT = PatternFill("solid", fgColor=TEAL_LIGHT)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", indent=1)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP   = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

THIN_BORDER = Border(
    bottom=Side(style="thin", color=MID_GRAY)
)
HEADER_BORDER = Border(
    bottom=Side(style="medium", color=NAVY)
)


def _cur_fmt(currency):
    if currency == "CAD":
        return '"CA$"#,##0.00'
    return '$#,##0.00'


def _pct_fmt():
    return '0.0"%"'


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _write_header_row(ws, row, headers, start_col=2):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = HEADER_BORDER
    return row


def _write_data_row(ws, row, values, start_col=2, formats=None, is_alt=False):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = FONT_BODY
        c.border = THIN_BORDER
        if is_alt:
            c.fill = FILL_ALT
        if formats and i < len(formats) and formats[i]:
            c.number_format = formats[i]
            c.alignment = ALIGN_RIGHT
        elif isinstance(v, (int, float)):
            c.alignment = ALIGN_RIGHT
        else:
            c.alignment = ALIGN_LEFT
    return row


def _write_kpi(ws, row, col, value, label, fmt=None):
    vc = ws.cell(row=row, column=col, value=value)
    vc.font = FONT_KPI_VAL
    vc.alignment = Alignment(horizontal="center", vertical="bottom")
    if fmt:
        vc.number_format = fmt
    lc = ws.cell(row=row + 1, column=col, value=label)
    lc.font = FONT_KPI_LBL
    lc.alignment = Alignment(horizontal="center", vertical="top")


def _write_section_title(ws, row, col, title, span=6):
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + span - 1
    )
    c = ws.cell(row=row, column=col, value=title)
    c.font = FONT_H2
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(bottom=Side(style="medium", color=TEAL))
    for sc in range(col + 1, col + span):
        ws.cell(row=row, column=sc).border = Border(
            bottom=Side(style="medium", color=TEAL)
        )


def _truncate_sheet_name(name, existing_names, max_len=31):
    """Excel sheet names max 31 chars. Remove illegal chars. Ensure unique."""
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '')
    name = name.strip()
    if len(name) > max_len:
        name = name[:max_len - 2] + '..'
    # Ensure uniqueness
    base = name
    counter = 2
    while name in existing_names:
        suffix = f' {counter}'
        name = base[:max_len - len(suffix)] + suffix
        counter += 1
    existing_names.add(name)
    return name


def _generate_summary_only_excel(wb, results, company_name, cur, cf, is_admin):
    """Fallback for old-format analyses that have no shipments array.
    Generates a summary workbook from by_service, by_carrier, and totals."""
    by_service = results.get("by_service", {})
    by_carrier = results.get("by_carrier", {})
    total_current = results.get("total_current", 0)
    total_new = results.get("total_new", 0)
    total_savings = results.get("total_savings", 0)
    savings_pct = (total_savings / total_current * 100) if total_current else 0

    ws = wb.active
    ws.title = "Analysis Summary"
    ws.sheet_properties.tabColor = NAVY
    ws.column_dimensions['A'].width = 3

    # Title banner
    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = FILL_NAVY
        ws.cell(row=2, column=col).fill = FILL_NAVY
        ws.cell(row=3, column=col).fill = FILL_NAVY
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 24

    ws.merge_cells('B2:I2')
    ws.cell(row=2, column=2, value=f"Shipping Analysis – {company_name}").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells('B3:I3')
    ws.cell(row=3, column=2,
            value=f"Prepared by Broad Reach Logistics  •  {datetime.now().strftime('%B %d, %Y')}").font = Font(
        name="Aptos", size=10, color="94A3B8")
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="left", vertical="center")

    # KPIs
    r = 5
    ws.row_dimensions[r].height = 42
    ws.row_dimensions[r + 1].height = 18
    kpis = [
        (2, total_current, "Current Total Spend", cf),
        (4, total_new, "Broad Reach Price", cf),
        (6, total_savings, "Total Savings", cf),
        (8, savings_pct / 100 if savings_pct else 0, "Savings Rate", '0.0%'),
    ]
    for col, val, lbl, fmt in kpis:
        for kr in (r, r + 1):
            ws.cell(row=kr, column=col).fill = FILL_TEAL_LT
            ws.cell(row=kr, column=col + 1).fill = FILL_TEAL_LT
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        _write_kpi(ws, r, col, val, lbl, fmt)

    # By Service table
    r = 8
    if by_service:
        _write_section_title(ws, r, 2, "Breakdown by Service", span=6)
        r += 1
        _write_header_row(ws, r, ["Service", "Current Spend", "BR Price", "Savings", "Savings %"], start_col=2)
        r += 1
        for idx, (svc, data) in enumerate(sorted(by_service.items(),
                key=lambda x: x[1] if isinstance(x[1], (int, float)) else x[1].get("savings", 0), reverse=True)):
            if isinstance(data, (int, float)):
                vals = [svc, data, "", "", ""]
                fmts = [None, cf, cf, cf, _pct_fmt()]
            else:
                vals = [svc, data.get("original", 0), data.get("br", 0),
                        data.get("savings", 0), data.get("savings_pct", 0)]
                fmts = [None, cf, cf, cf, _pct_fmt()]
            _write_data_row(ws, r, vals, start_col=2, formats=fmts, is_alt=(idx % 2 == 1))
            r += 1
        r += 1

    # By Carrier table
    if by_carrier:
        _write_section_title(ws, r, 2, "Breakdown by Carrier", span=6)
        r += 1
        _write_header_row(ws, r, ["Carrier", "Current Spend", "BR Price", "Savings", "Savings %"], start_col=2)
        r += 1
        for idx, (carr, data) in enumerate(sorted(by_carrier.items(),
                key=lambda x: x[1] if isinstance(x[1], (int, float)) else x[1].get("savings", 0), reverse=True)):
            if isinstance(data, (int, float)):
                vals = [carr, data, "", "", ""]
                fmts = [None, cf, cf, cf, _pct_fmt()]
            else:
                vals = [carr, data.get("original", 0), data.get("br", 0),
                        data.get("savings", 0), data.get("savings_pct", 0)]
                fmts = [None, cf, cf, cf, _pct_fmt()]
            _write_data_row(ws, r, vals, start_col=2, formats=fmts, is_alt=(idx % 2 == 1))
            r += 1

    widths = {2: 30, 3: 18, 4: 18, 5: 16, 6: 14, 7: 14, 8: 16, 9: 14}
    for col, w in widths.items():
        _set_col_width(ws, col, w)

    for ws_iter in wb.worksheets:
        ws_iter.page_setup.orientation = "landscape"
        ws_iter.page_setup.fitToWidth = 1
        ws_iter.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_analysis_excel(results, company_name, currency=None, role="client"):
    """
    Generate a branded Excel workbook from analysis results.
    role='client' strips internal cost/profit fields.
    role='admin' includes buy_price, profit, margin columns.
    Returns bytes of the .xlsx file.
    """
    wb = Workbook()
    cur = currency or results.get("currency", "USD")
    cf = _cur_fmt(cur)
    summary = results.get("summary", {})
    shipments = results.get("shipments", [])
    by_service = results.get("by_service", {})
    by_zone = results.get("by_zone", {})
    by_carrier = results.get("by_carrier", {})
    service_mix = results.get("br_service_mix", {})
    is_admin = (role != "client")

    # If old-format analysis (no shipments), generate summary-only workbook
    if not shipments:
        return _generate_summary_only_excel(wb, results, company_name, cur, cf, is_admin)

    # Collect all unique rated service names
    all_service_names = set()
    for s in shipments:
        if s.get("all_rates"):
            all_service_names.update(s["all_rates"].keys())
    service_list = sorted(all_service_names)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 – Executive Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = NAVY
    ws.column_dimensions['A'].width = 3

    # Title banner
    for col in range(1, 12):
        ws.cell(row=1, column=col).fill = FILL_NAVY
        ws.cell(row=2, column=col).fill = FILL_NAVY
        ws.cell(row=3, column=col).fill = FILL_NAVY
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 24

    ws.merge_cells('B2:K2')
    tc = ws.cell(row=2, column=2, value=f"Shipping Analysis – {company_name}")
    tc.font = FONT_TITLE
    tc.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells('B3:K3')
    sc = ws.cell(row=3, column=2,
                 value=f"Prepared by Broad Reach Logistics  •  {datetime.now().strftime('%B %d, %Y')}")
    sc.font = Font(name="Aptos", size=10, color="94A3B8")
    sc.alignment = Alignment(horizontal="left", vertical="center")

    # KPI row
    r = 5
    ws.row_dimensions[r].height = 42
    ws.row_dimensions[r + 1].height = 18

    total_orig = summary.get("total_original", 0)
    total_br = summary.get("total_br", 0)
    total_savings = summary.get("total_savings", 0)
    savings_pct = summary.get("savings_pct", 0)
    ship_count = summary.get("shipment_count", len(shipments))
    ships_w_savings = summary.get("shipments_with_savings", 0)

    # KPI cells with background tint
    kpi_data = [
        (2, ship_count, "Shipments Analyzed", '#,##0'),
        (4, total_orig, "Current Total Spend", cf),
        (6, total_br, "Broad Reach Price", cf),
        (8, total_savings, "Total Savings", cf),
        (10, savings_pct / 100 if savings_pct else 0, "Savings Rate", '0.0%'),
    ]
    for col, val, lbl, fmt in kpi_data:
        for kr in (r, r + 1):
            ws.cell(row=kr, column=col).fill = FILL_TEAL_LT
            ws.cell(row=kr, column=col + 1).fill = FILL_TEAL_LT
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        _write_kpi(ws, r, col, val, lbl, fmt)

    # Annualized savings
    r = 8
    _write_section_title(ws, r, 2, "Projected Annual Impact (250 Business Days)", span=10)
    r += 1
    ws.row_dimensions[r].height = 36
    ws.row_dimensions[r + 1].height = 18

    if ship_count > 0:
        # Calculate annualization factor from shipment date range (250 biz days/yr)
        ship_dates = sorted([s.get("ship_date", "") for s in shipments if s.get("ship_date")])
        if len(ship_dates) > 1:
            try:
                from datetime import datetime as _dt
                first = _dt.strptime(ship_dates[0][:10], "%Y-%m-%d")
                last  = _dt.strptime(ship_dates[-1][:10], "%Y-%m-%d")
                cal_days = max(1, (last - first).days + 1)
                biz_days_in_range = max(1, round(cal_days * 5 / 7))
                annual_factor = 250 / biz_days_in_range
            except Exception:
                annual_factor = 12
        else:
            annual_factor = 12  # single day sample → multiply by 12 as fallback
        ann_savings = total_savings * annual_factor
        per_ship_savings = total_savings / max(ship_count, 1)
    else:
        ann_shipments = 0
        ann_savings = 0
        per_ship_savings = 0

    ann_kpis = [
        (2, per_ship_savings, "Avg Savings / Shipment", cf),
        (4, total_savings, f"Sample Savings ({ship_count} shipments)", cf),
        (6, ann_savings, "Projected Annual Savings", cf),
        (8, ships_w_savings, "Shipments With Savings", '#,##0'),
    ]
    for col, val, lbl, fmt in ann_kpis:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=col + 1)
        _write_kpi(ws, r, col, val, lbl, fmt)

    # Savings by Current Service table
    r = 12
    _write_section_title(ws, r, 2, "Savings Breakdown by Original Service", span=6)
    r += 1
    headers_svc = ["Service", "Shipments", "Current Spend", "Broad Reach Price", "Savings", "Savings %"]
    _write_header_row(ws, r, headers_svc, start_col=2)
    r += 1
    for idx, (svc, data) in enumerate(sorted(by_service.items(), key=lambda x: x[1].get("savings", 0), reverse=True)):
        vals = [
            svc,
            data.get("count", 0),
            data.get("original", 0),
            data.get("br", 0),
            data.get("savings", 0),
            data.get("savings_pct", 0),
        ]
        fmts = [None, '#,##0', cf, cf, cf, _pct_fmt()]
        _write_data_row(ws, r, vals, start_col=2, formats=fmts, is_alt=(idx % 2 == 1))
        r += 1

    # Totals row
    total_row_vals = [
        "TOTAL", ship_count, total_orig, total_br, total_savings, savings_pct
    ]
    for i, v in enumerate(total_row_vals):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        c.fill = FILL_TEAL
        c.alignment = ALIGN_RIGHT if i > 0 else ALIGN_LEFT
        if i == 2 or i == 3 or i == 4:
            c.number_format = cf
        elif i == 5:
            c.number_format = _pct_fmt()
        elif i == 1:
            c.number_format = '#,##0'
    r += 2

    # Savings by Zone
    if by_zone:
        _write_section_title(ws, r, 2, "Savings by Zone", span=6)
        r += 1
        zone_headers = ["Zone", "Shipments", "Avg Current", "Avg BR Price", "Savings", "Savings %"]
        _write_header_row(ws, r, zone_headers, start_col=2)
        r += 1
        for idx, (zone, data) in enumerate(sorted(by_zone.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999)):
            vals = [
                f"Zone {zone}",
                data.get("count", 0),
                data.get("avg_original", 0),
                data.get("avg_br", 0),
                data.get("savings", 0),
                data.get("savings_pct", 0),
            ]
            fmts = [None, '#,##0', cf, cf, cf, _pct_fmt()]
            _write_data_row(ws, r, vals, start_col=2, formats=fmts, is_alt=(idx % 2 == 1))
            r += 1
        r += 1

    # Set column widths for summary
    widths = {2: 28, 3: 14, 4: 18, 5: 18, 6: 16, 7: 14, 8: 18, 9: 18, 10: 16, 11: 14}
    for col, w in widths.items():
        _set_col_width(ws, col, w)

    # Chart – Savings by Service
    if by_service and len(by_service) > 1:
        chart_start = r
        _write_section_title(ws, r, 2, "Savings Distribution by Service", span=6)
        r += 1
        # Write chart data in hidden helper area
        chart_data_row = r
        ws.cell(row=r, column=2, value="Service").font = FONT_BODY_SM
        ws.cell(row=r, column=3, value="Savings").font = FONT_BODY_SM
        r += 1
        for svc, data in sorted(by_service.items(), key=lambda x: x[1].get("savings", 0), reverse=True):
            ws.cell(row=r, column=2, value=svc).font = FONT_BODY_SM
            ws.cell(row=r, column=3, value=data.get("savings", 0)).font = FONT_BODY_SM
            ws.cell(row=r, column=3).number_format = cf
            r += 1

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = None
        chart.y_axis.title = f"Savings ({cur})"
        chart.x_axis.title = None
        chart.legend = None
        data_ref = Reference(ws, min_col=3, min_row=chart_data_row, max_row=r - 1)
        cats_ref = Reference(ws, min_col=2, min_row=chart_data_row + 1, max_row=r - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 20
        chart.height = 10
        # Color bars teal
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = TEAL
        ws.add_chart(chart, f"E{chart_start}")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 – Shipment Detail (Lowest Cost)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Lowest Cost Comparison")
    ws2.sheet_properties.tabColor = TEAL
    ws2.column_dimensions['A'].width = 3

    # Title
    for col in range(1, 20):
        ws2.cell(row=1, column=col).fill = FILL_NAVY
    ws2.row_dimensions[1].height = 32
    ws2.merge_cells('B1:S1')
    t2 = ws2.cell(row=1, column=2, value="Shipment-by-Shipment Comparison – Lowest Cost Match")
    t2.font = Font(name="Aptos", size=14, bold=True, color=WHITE)
    t2.alignment = Alignment(horizontal="left", vertical="center")

    r2 = 3
    detail_headers = [
        "Ship Date", "Tracking", "Carrier", "Service",
        "Weight", "Billed Wt", "Dims (L×W×H)",
        "Origin", "Destination", "Zone",
        f"Current Price ({cur})", "Best BR Service",
        f"BR Price ({cur})"
    ]
    # If original price exists, add savings columns
    has_original_price = any(s.get("price", 0) > 0 for s in shipments)
    if has_original_price:
        detail_headers += [f"Savings ({cur})", "Savings %"]
    # Admin-only columns
    if is_admin:
        detail_headers += [f"Buy Price ({cur})", f"Profit ({cur})", "Margin %"]

    _write_header_row(ws2, r2, detail_headers, start_col=2)
    r2 += 1
    ws2.freeze_panes = f"A{r2}"

    for idx, s in enumerate(shipments):
        dims = ""
        if s.get("length") and s.get("width") and s.get("height"):
            dims = f'{s["length"]}×{s["width"]}×{s["height"]}'
        origin = f'{s.get("origin_zip", "")} {s.get("origin_state", "")}'.strip()
        dest = f'{s.get("dest_zip", "")} {s.get("dest_state", "")}'.strip()

        vals = [
            s.get("ship_date", ""),
            s.get("tracking", ""),
            s.get("carrier", ""),
            s.get("service", ""),
            s.get("weight", ""),
            s.get("billable_weight") or s.get("billed_weight", ""),
            dims,
            origin,
            dest,
            s.get("zone", ""),
            s.get("price", 0),
            s.get("br_service", ""),
            s.get("br_price", 0),
        ]
        fmts = [
            None, None, None, None,
            '0.0', '0.0', None,
            None, None, '#,##0',
            cf, None, cf,
        ]
        if has_original_price:
            vals += [s.get("savings", 0), s.get("savings_pct", 0)]
            fmts += [cf, _pct_fmt()]
        if is_admin:
            # Find the best rate's buy_price, profit, margin from all_rates for br_service
            br_svc = s.get("br_service", "")
            best_rate = (s.get("all_rates") or {}).get(br_svc, {})
            buy = best_rate.get("buy_price", s.get("buy_price", 0)) or 0
            profit = best_rate.get("profit", s.get("profit", 0)) or 0
            margin = best_rate.get("margin_pct", s.get("margin_pct", 0)) or 0
            vals += [buy, profit, margin]
            fmts += [cf, cf, _pct_fmt()]

        _write_data_row(ws2, r2, vals, start_col=2, formats=fmts, is_alt=(idx % 2 == 1))

        # Color-code savings
        if has_original_price:
            sav_offset = (len(detail_headers) - (3 if is_admin else 0)) - 2
            sav_cell = ws2.cell(row=r2, column=2 + sav_offset)
            pct_cell = ws2.cell(row=r2, column=2 + sav_offset + 1)
            if (s.get("savings", 0) or 0) > 0:
                sav_cell.font = FONT_GREEN
                pct_cell.font = FONT_GREEN
            elif (s.get("savings", 0) or 0) < 0:
                sav_cell.font = FONT_RED
                pct_cell.font = FONT_RED
        # Color-code profit for admin
        if is_admin:
            profit_cell = ws2.cell(row=r2, column=2 + len(detail_headers) - 2)
            if profit_cell.value and profit_cell.value > 0:
                profit_cell.font = FONT_GREEN
            elif profit_cell.value and profit_cell.value < 0:
                profit_cell.font = FONT_RED
        r2 += 1

    # Totals row
    if shipments:
        r2_total = r2
        totals = ["", "", "", "TOTALS", "", "", "", "", "", "",
                   total_orig, "", total_br]
        if has_original_price:
            totals += [total_savings, savings_pct]
        if is_admin:
            # Sum buy/profit across all shipments
            total_buy = 0
            total_profit = 0
            for s in shipments:
                br_svc = s.get("br_service", "")
                best_rate = (s.get("all_rates") or {}).get(br_svc, {})
                total_buy += best_rate.get("buy_price", s.get("buy_price", 0)) or 0
                total_profit += best_rate.get("profit", s.get("profit", 0)) or 0
            total_margin = (total_profit / total_br * 100) if total_br else 0
            totals += [total_buy, total_profit, total_margin]
        fmts_t = [None]*10 + [cf, None, cf]
        if has_original_price:
            fmts_t += [cf, _pct_fmt()]
        if is_admin:
            fmts_t += [cf, cf, _pct_fmt()]
        for i, v in enumerate(totals):
            c = ws2.cell(row=r2, column=2 + i, value=v)
            c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
            c.fill = FILL_TEAL
            c.alignment = ALIGN_RIGHT if i >= 10 else ALIGN_LEFT
            if i < len(fmts_t) and fmts_t[i]:
                c.number_format = fmts_t[i]

    # Column widths
    detail_widths = [12, 18, 10, 22, 8, 9, 14, 14, 14, 7, 16, 30, 16]
    if has_original_price:
        detail_widths += [14, 12]
    if is_admin:
        detail_widths += [14, 14, 10]
    for i, w in enumerate(detail_widths):
        _set_col_width(ws2, 2 + i, w)

    # Add data bars on savings column
    if has_original_price and len(shipments) > 1:
        sav_col_idx = (len(detail_headers) - (3 if is_admin else 0)) - 2
        sav_col_letter = get_column_letter(2 + sav_col_idx)
        rule = DataBarRule(start_type='min', end_type='max', color=TEAL)
        ws2.conditional_formatting.add(
            f'{sav_col_letter}4:{sav_col_letter}{r2 - 1}', rule
        )

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3+ – One tab per rated service (rate card view)
    # ══════════════════════════════════════════════════════════════════════════
    used_sheet_names = {"Executive Summary", "Lowest Cost Comparison", "All Services Matrix"}
    for svc_name in service_list:
        sheet_name = _truncate_sheet_name(svc_name, used_sheet_names)
        ws_svc = wb.create_sheet(sheet_name)
        ws_svc.sheet_properties.tabColor = "334155"
        ws_svc.column_dimensions['A'].width = 3

        # Header
        for col in range(1, 14):
            ws_svc.cell(row=1, column=col).fill = FILL_NAVY
        ws_svc.row_dimensions[1].height = 32
        ws_svc.merge_cells('B1:M1')
        ws_svc.cell(row=1, column=2, value=svc_name).font = Font(
            name="Aptos", size=14, bold=True, color=WHITE
        )
        ws_svc.cell(row=1, column=2).alignment = Alignment(horizontal="left", vertical="center")

        rs = 3
        svc_headers = [
            "#", "Ship Date", "Tracking",
            "Weight", "Billed Wt", "Zone",
            "Origin", "Destination",
            f"Service Price ({cur})",
        ]
        if has_original_price:
            svc_headers += [f"Current Price ({cur})", f"Savings ({cur})", "Savings %"]

        # Add fuel/accessorial columns if data has them
        sample_rate = None
        for s in shipments:
            if s.get("all_rates", {}).get(svc_name):
                sample_rate = s["all_rates"][svc_name]
                break
        has_fuel = sample_rate and "fuel" in sample_rate
        has_access = sample_rate and "accessorials" in sample_rate

        if has_fuel:
            svc_headers.insert(-3 if has_original_price else -0, f"Fuel ({cur})")
        if has_access:
            ins_pos = (-3 if has_original_price else len(svc_headers))
            if has_fuel:
                ins_pos = ins_pos  # after fuel
            svc_headers.insert(ins_pos, f"Accessorials ({cur})")

        # Rebuild properly - simplify
        svc_headers = ["#", "Ship Date", "Tracking", "Weight", "Billed Wt", "Zone",
                        "Origin", "Destination"]
        svc_fmts = [None, None, None, '0.0', '0.0', '#,##0', None, None]

        svc_headers.append(f"Price ({cur})")
        svc_fmts.append(cf)

        if has_fuel:
            svc_headers.append(f"Fuel ({cur})")
            svc_fmts.append(cf)
        if has_access:
            svc_headers.append(f"Accessorials ({cur})")
            svc_fmts.append(cf)
        if has_original_price:
            svc_headers += [f"Your Current ({cur})", f"Difference ({cur})", "Diff %"]
            svc_fmts += [cf, cf, _pct_fmt()]
        if is_admin:
            svc_headers += [f"Buy Price ({cur})", f"Profit ({cur})", "Margin %"]
            svc_fmts += [cf, cf, _pct_fmt()]

        _write_header_row(ws_svc, rs, svc_headers, start_col=2)
        rs += 1
        ws_svc.freeze_panes = f"A{rs}"

        svc_total_price = 0
        svc_total_current = 0
        svc_total_buy = 0
        svc_total_profit = 0
        svc_count = 0

        for idx, s in enumerate(shipments):
            rate = (s.get("all_rates") or {}).get(svc_name)
            if not rate:
                continue
            svc_count += 1
            origin = f'{s.get("origin_zip", "")} {s.get("origin_state", "")}'.strip()
            dest = f'{s.get("dest_zip", "")} {s.get("dest_state", "")}'.strip()

            price = rate.get("final", rate.get("sell_price", 0))
            svc_total_price += price

            vals = [svc_count, s.get("ship_date", ""), s.get("tracking", ""),
                    s.get("weight", ""), rate.get("billable_wt", s.get("billable_weight", "")),
                    rate.get("zone", s.get("zone", "")),
                    origin, dest, price]

            if has_fuel:
                vals.append(rate.get("fuel", 0))
            if has_access:
                vals.append(rate.get("accessorials", 0))
            if has_original_price:
                cur_price = s.get("price", 0)
                svc_total_current += cur_price
                diff = cur_price - price
                diff_pct = (diff / cur_price * 100) if cur_price else 0
                vals += [cur_price, diff, diff_pct]
            if is_admin:
                buy = rate.get("buy_price", 0) or 0
                profit = rate.get("profit", 0) or 0
                margin = rate.get("margin_pct", 0) or 0
                svc_total_buy += buy
                svc_total_profit += profit
                vals += [buy, profit, margin]

            _write_data_row(ws_svc, rs, vals, start_col=2, formats=svc_fmts,
                           is_alt=((svc_count - 1) % 2 == 1))

            # Color savings/losses
            if has_original_price:
                admin_offset = 3 if is_admin else 0
                diff_cell = ws_svc.cell(row=rs, column=2 + len(svc_headers) - 2 - admin_offset)
                pct_cell = ws_svc.cell(row=rs, column=2 + len(svc_headers) - 1 - admin_offset)
                diff_val = diff if has_original_price else 0
                if diff_val > 0:
                    diff_cell.font = FONT_GREEN
                    pct_cell.font = FONT_GREEN
                elif diff_val < 0:
                    diff_cell.font = FONT_RED
                    pct_cell.font = FONT_RED
            if is_admin:
                profit_cell = ws_svc.cell(row=rs, column=2 + len(svc_headers) - 2)
                if profit_cell.value and profit_cell.value > 0:
                    profit_cell.font = FONT_GREEN
                elif profit_cell.value and profit_cell.value < 0:
                    profit_cell.font = FONT_RED
            rs += 1

        # Totals row for service tab
        if svc_count > 0:
            tot_vals = ["", "", "", "", "", "", "", "TOTALS", svc_total_price]
            if has_fuel:
                tot_vals.append("")
            if has_access:
                tot_vals.append("")
            if has_original_price:
                diff_tot = svc_total_current - svc_total_price
                diff_pct_tot = (diff_tot / svc_total_current * 100) if svc_total_current else 0
                tot_vals += [svc_total_current, diff_tot, diff_pct_tot]
            if is_admin:
                svc_margin_tot = (svc_total_profit / svc_total_price * 100) if svc_total_price else 0
                tot_vals += [svc_total_buy, svc_total_profit, svc_margin_tot]

            for i, v in enumerate(tot_vals):
                c = ws_svc.cell(row=rs, column=2 + i, value=v)
                c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
                c.fill = PatternFill("solid", fgColor="334155")
                c.alignment = ALIGN_RIGHT if i >= 8 else ALIGN_LEFT
                if i < len(svc_fmts) and svc_fmts[i]:
                    c.number_format = svc_fmts[i]

        # Column widths
        for i in range(len(svc_headers)):
            w = 14
            if i == 0:
                w = 5
            elif i == 1:
                w = 12
            elif i == 2:
                w = 18
            elif i in (3, 4):
                w = 10
            elif i == 5:
                w = 7
            elif i in (6, 7):
                w = 16
            _set_col_width(ws_svc, 2 + i, w)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB – All Services Comparison (wide view)
    # ══════════════════════════════════════════════════════════════════════════
    if service_list and shipments:
        ws_all = wb.create_sheet("All Services Matrix")
        ws_all.sheet_properties.tabColor = "7C3AED"
        ws_all.column_dimensions['A'].width = 3

        for col in range(1, 8 + len(service_list)):
            ws_all.cell(row=1, column=col).fill = FILL_NAVY
        ws_all.row_dimensions[1].height = 32
        ws_all.merge_cells(start_row=1, start_column=2,
                          end_row=1, end_column=7 + len(service_list))
        ws_all.cell(row=1, column=2,
                   value="All Services Price Matrix – Side-by-Side Comparison").font = Font(
            name="Aptos", size=14, bold=True, color=WHITE
        )
        ws_all.cell(row=1, column=2).alignment = Alignment(horizontal="left", vertical="center")

        ra = 3
        all_headers = ["#", "Tracking", "Weight", "Zone"]
        if has_original_price:
            all_headers.append(f"Current ({cur})")
        all_headers.append(f"Best BR ({cur})")
        for sn in service_list:
            # Truncate long service names for column headers
            short = sn[:25] + '..' if len(sn) > 27 else sn
            all_headers.append(short)

        _write_header_row(ws_all, ra, all_headers, start_col=2)
        ra += 1
        ws_all.freeze_panes = f"A{ra}"

        for idx, s in enumerate(shipments):
            vals = [
                idx + 1,
                s.get("tracking", ""),
                s.get("weight", ""),
                s.get("zone", ""),
            ]
            fmts_a = [None, None, '0.0', '#,##0']
            if has_original_price:
                vals.append(s.get("price", 0))
                fmts_a.append(cf)
            vals.append(s.get("br_price", 0))
            fmts_a.append(cf)

            for sn in service_list:
                rate = (s.get("all_rates") or {}).get(sn)
                if rate:
                    vals.append(rate.get("final", rate.get("sell_price", "")))
                else:
                    vals.append("")
                fmts_a.append(cf)

            _write_data_row(ws_all, ra, vals, start_col=2, formats=fmts_a,
                           is_alt=(idx % 2 == 1))

            # Highlight the lowest price cell in green
            svc_start_col = 2 + (6 if has_original_price else 5)
            min_price = None
            min_col = None
            for si, sn in enumerate(service_list):
                rate = (s.get("all_rates") or {}).get(sn)
                if rate:
                    p = rate.get("final", rate.get("sell_price", 999999))
                    if min_price is None or p < min_price:
                        min_price = p
                        min_col = svc_start_col + si
            if min_col:
                cell = ws_all.cell(row=ra, column=min_col)
                cell.fill = FILL_GREEN
                cell.font = FONT_GREEN

            ra += 1

        # Column widths
        _set_col_width(ws_all, 2, 5)   # #
        _set_col_width(ws_all, 3, 18)  # tracking
        _set_col_width(ws_all, 4, 8)   # weight
        _set_col_width(ws_all, 5, 7)   # zone
        base = 6
        if has_original_price:
            _set_col_width(ws_all, base, 16)
            base += 1
        _set_col_width(ws_all, base, 14)
        for si in range(len(service_list)):
            _set_col_width(ws_all, base + 1 + si, 16)

    # ══════════════════════════════════════════════════════════════════════════
    # Print settings for all sheets
    # ══════════════════════════════════════════════════════════════════════════
    for ws_iter in wb.worksheets:
        ws_iter.page_setup.orientation = "landscape"
        ws_iter.page_setup.fitToWidth = 1
        ws_iter.page_setup.fitToHeight = 0

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
